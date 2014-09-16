from __future__ import unicode_literals

from timelines.unicsv import UnicodeCSVWriter

from django.contrib.auth.decorators import permission_required, login_required
from django.core.urlresolvers import reverse
from django.http import HttpResponse, HttpResponseRedirect
from django.utils.decorators import method_decorator
from django.views.generic import View
from django.views.generic.base import TemplateView
from django.template import RequestContext
from django.shortcuts import render_to_response
from django.core.cache import cache

from django_tables2 import RequestConfig

from .forms import OccurrenceFilterForm
from .tables import ApptTable
from .models import SubscriptionView, PerformanceView
from openpyxl import Workbook
from openpyxl.cell import get_column_letter


class OccurrenceMixin(object):
    """Allow filtering by"""
    @method_decorator(permission_required('timelines.view_occurrence'))
    def dispatch(self, request, *args, **kwargs):
        self.form = OccurrenceFilterForm(request.GET)
        self.items = self.form.get_items()
        return super(OccurrenceMixin, self).dispatch(request, *args, **kwargs)


class OccurrenceList(OccurrenceMixin, TemplateView):
    """Displays a paginated list of occurrences."""
    template_name = 'timelines/timelines_list.html'
    table_template_name = 'django_tables2/bootstrap-tables.html'
    items_per_page = 10

    def get_table(self):
        table = ApptTable(self.items, template=self.table_template_name)
        paginate = {'per_page': self.items_per_page}
        RequestConfig(self.request, paginate=paginate).configure(table)
        return table

    def get_context_data(self, *args, **kwargs):
        return {
            'form': self.form,
            'table': self.get_table()
        }


class CSVOccurrenceList(OccurrenceMixin, View):
    """Export filtered reports to a CSV file."""
    # Fields to include in the csv, in order.
    filename = 'occurrences'

    def get_table(self):
        table = ApptTable(self.items)
        RequestConfig(self.request).configure(table)
        return table

    def get(self, request, *args, **kwargs):
        if not self.form.is_valid():
            url = reverse('occurrence_list')
            if request.GET:
                url = '{0}?{1}'.format(url, request.GET.urlencode())
            return HttpResponseRedirect(url)

        response = HttpResponse(content_type='text/csv')
        content_disposition = 'attachment; filename=%s.csv' % self.filename
        response['Content-Disposition'] = content_disposition
        writer = UnicodeCSVWriter(response)
        writer.writerows(self.get_data())
        return response

    def get_data(self):
        table = self.get_table()
        columns = [x.title() for x in table.columns.names()]
        rows = [columns, ]
        for item in table.rows:
            cells = [x for x in item]
            row = []
            for cell in cells:
                row.append(cell)
            rows.append(row)
        return rows


@login_required
def subscriptionsView(request):
    if not cache.get('subs'):
        subs = SubscriptionView.objects.all()
        cache.set('subs', subs, 60 * 10)
    else:
        subs = cache.get('subs')
    return render_to_response(
        "timelines/subscriptions.html",
        {'subs': subs},
        context_instance=RequestContext(request))


@login_required
def performanceView(request):
    if not cache.get('performance'):
        performance = PerformanceView.objects.all().order_by(
            '-advice_subs', '-preg_subs', '-birth_subs', '-cvisits')
        cache.set('performance', performance, 60 * 10)
    else:
        performance = cache.get('performance')
    return render_to_response(
        "timelines/performance.html",
        {'performance': performance},
        context_instance=RequestContext(request))


@login_required
def performanceExcelView(request):
    wb = Workbook(encoding='utf-8')
    sheet1 = wb.worksheets[0]
    header = [
        'Name', 'Phone Number', 'Creation Date', 'Facility', 'Village',
        'ANC/PNC Advice Registrations', 'Preg/ANC Visit Registered',
        'Birth/PNC Visits Registered', 'No. of Confirmed Visits', 'Last Reporting Date'
    ]
    if not cache.get('performance'):
        performance = PerformanceView.objects.all().order_by(
            '-advice_subs', '-preg_subs', '-birth_subs', '-cvisits'
        )
    else:
        performance = cache.get('performance')
    performance_data = performance.values_list(
        'name', 'identity', 'created_on', 'facility', 'village',
        'advice_subs', 'preg_subs', 'birth_subs', 'cvisits',
        'last_reporting_date'
    )

    for hcol, hcol_data in enumerate(header, start=1):
        col_idx = get_column_letter(hcol)
        sheet1.cell('%s%s' % (col_idx, 1)).value = hcol_data

    for row, row_data in enumerate(performance_data, start=2):  # start from row no.2
        for col, col_data in enumerate(row_data, start=1):
            col_idx = get_column_letter(col)
            sheet1.cell('%s%s' % (col_idx, row)).value = col_data
    response = HttpResponse(mimetype='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=performance_data.xlsx'
    wb.save(response)
    return response
